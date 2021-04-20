'''
from openpyxl import Workbook

wb = Workbook()
#print(type(wb))

# designe la feuille active
ws = wb.active

# insertion dans les cellules ciblées
ws['A1'] = 42

import datetime
ws['A2'] = datetime.datetime.now()

# enregistrement du fichier
wb.save('files/sample.xlsx')
'''

import openpyxl as xl

wb  = xl.load_workbook('files/students.xlsx')
ws = wb['notes']

#print(type(ws)) # class Worksheet

for row in range(2, 6):
  n1 = ws.cell(row, 2).value
  n2 = ws.cell(row, 3).value
  n3 = ws.cell(row, 4).value
  s = n1+n2+n3
  ws.cell(row, 5).value = s/3
  
#ws.cell(1,2).value = 'test' # écture en B1 (Row = 1, Col = 2)

# production d'un graphique et insertion en E10
from openpyxl.chart import BarChart, Reference
values = Reference(ws, min_row=2, max_row=2, min_col=2, max_col=4)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, 'E10')

# enregistrement du fichier
wb.save('files/new_students.xlsx')